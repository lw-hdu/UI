'''
Descripttion: 
Author: Liuwen
Date: 2022-02-28 14:23:50
LastEditTime: 2022-03-03 15:52:57
'''
import os
import openpyxl

class ReadExcel:
    def get_file_path(self):
        return os.path.abspath(__file__).split('common')[0]

    def get_xls(self):
        wb = openpyxl.load_workbook(self.get_file_path()+'data\login_data.xlsx')
        sheet = wb['login']
        # print(sheet.max_column,sheet.max_row)
        all_list = []
        for row in range(2,sheet.max_row+1):
            temp_list = []
            for colum in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(row,colum).value)
            all_list.append(temp_list)
        return all_list
        #print(all_list)
            
# if __name__ == '__main__':
#     ReadExcel().get_xls()